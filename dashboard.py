import dash
from dash import dcc, html, Input, Output, dash_table
import plotly.graph_objs as go
import pandas as pd
from openpyxl import load_workbook

# Import helper functions from process_excel.py
from process_excel import safe_float, find_table_boundaries

# Initialize Dash app
app = dash.Dash(__name__)
app.title = "Traffic Insight Dashboard"

# Color scheme
colors = {
    'background': '#f8f9fa',
    'text': '#2c3e50',
    'primary': '#3498db',
    'secondary': '#2ecc71',
    'accent': '#e74c3c',
    'card': '#ffffff'
}

def load_excel_data():
    """Load and parse data from Excel file"""
    try:
        wb = load_workbook('keywords.xlsx')
        ws = wb['Traffic-Status']
        
        # Find all tables
        tables = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
            cell_a = row[0]
            if cell_a.value and isinstance(cell_a.value, str):
                if row[1].value == 'Month' or (row_idx < ws.max_row and ws.cell(row_idx+1, 2).value == 'Month'):
                    tables.append({
                        'title': cell_a.value,
                        'row_idx': row_idx
                    })
        
        # Extract data for each table
        table_data_list = []
        for table_info in tables:
            title = table_info['title']
            start_row = table_info['row_idx']
            
            boundaries = find_table_boundaries(ws, start_row)
            data_start = boundaries['data_start']
            data_end = boundaries['data_end']
            total_row = boundaries['total_row']
            
            # Extract monthly data
            monthly_data = []
            for row in range(data_start, data_end + 1):
                month = ws.cell(row, 2).value
                
                # Skip header rows or rows with non-month values
                if month is None or str(month).strip() in ['Month', 'Total', '% Change', '']:
                    continue
                
                val_2023 = safe_float(ws.cell(row, 3).value)
                val_2024 = safe_float(ws.cell(row, 4).value)
                val_2025 = safe_float(ws.cell(row, 5).value)
                yoy = ws.cell(row, 6).value
                lm = ws.cell(row, 7).value
                
                # Extract summary from column H
                summary_line = ws.cell(row, 8).value
                
                # Safely convert yoy and lm to float (skip if it's a header string)
                yoy_float = None
                if yoy is not None:
                    try:
                        # Skip if it looks like a header (contains letters that aren't part of a number)
                        if isinstance(yoy, str):
                            # Check if it's a header like "YOY % (2024-2025)"
                            if any(char.isalpha() for char in yoy.replace('.', '').replace('-', '').replace('(', '').replace(')', '').replace('%', '').strip()):
                                yoy_float = None
                            else:
                                yoy_float = float(yoy) * 100
                        elif isinstance(yoy, (int, float)):
                            yoy_float = float(yoy) * 100
                    except (ValueError, TypeError):
                        yoy_float = None
                
                lm_float = None
                if lm is not None:
                    try:
                        if isinstance(lm, str):
                            if any(char.isalpha() for char in lm.replace('.', '').replace('-', '').replace('(', '').replace(')', '').replace('%', '').strip()):
                                lm_float = None
                            else:
                                lm_float = float(lm) * 100
                        elif isinstance(lm, (int, float)):
                            lm_float = float(lm) * 100
                    except (ValueError, TypeError):
                        lm_float = None
                
                monthly_data.append({
                    'month': str(month) if month else '',
                    '2023': val_2023,
                    '2024': val_2024,
                    '2025': val_2025,
                    'yoy': yoy_float,
                    'lm': lm_float,
                    'summary': str(summary_line) if summary_line else ''
                })
            
            # Get totals
            total_2023 = safe_float(ws.cell(total_row, 3).value)
            total_2024 = safe_float(ws.cell(total_row, 4).value)
            total_2025 = safe_float(ws.cell(total_row, 5).value)
            yoy_total_cell = ws.cell(total_row, 6).value
            
            # Safely convert yoy_total to float (skip if it's a header string)
            yoy_total = None
            if yoy_total_cell is not None:
                try:
                    if isinstance(yoy_total_cell, str):
                        # Check if it's a header like "YOY % (2024-2025)"
                        if any(char.isalpha() for char in yoy_total_cell.replace('.', '').replace('-', '').replace('(', '').replace(')', '').replace('%', '').strip()):
                            yoy_total = None
                        else:
                            yoy_total = float(yoy_total_cell) * 100
                    elif isinstance(yoy_total_cell, (int, float)):
                        yoy_total = float(yoy_total_cell) * 100
                except (ValueError, TypeError):
                    yoy_total = None
            
            # Get full summary (all lines from column H)
            summary_lines = []
            for row in range(data_start, min(data_start + 15, ws.max_row + 1)):
                summary_val = ws.cell(row, 8).value
                if summary_val:
                    summary_lines.append(str(summary_val))
            
            table_data_list.append({
                'title': title,
                'monthly_data': monthly_data,
                'totals': {
                    '2023': total_2023,
                    '2024': total_2024,
                    '2025': total_2025,
                    'yoy_total': yoy_total
                },
                'summary': '\n'.join(summary_lines)
            })
        
        return table_data_list
    except Exception as e:
        print(f"Error loading data: {e}")
        return []

# Load data
data = load_excel_data()

# App layout
app.layout = html.Div([
    # Header
    html.Div([
        html.H1("Traffic Insight Dashboard", 
                style={'color': colors['text'], 'margin': '0', 'fontSize': '2.5rem'}),
        html.P("Interactive Analytics & AI-Powered Insights",
               style={'color': colors['text'], 'opacity': 0.7, 'margin': '10px 0 0 0'})
    ], style={'textAlign': 'center', 'padding': '20px', 'backgroundColor': colors['card'],
              'marginBottom': '20px', 'borderRadius': '10px', 'boxShadow': '0 2px 4px rgba(0,0,0,0.1)'}),
    
    # Table selector
    html.Div([
        html.Label("Select Table:", style={'fontWeight': 'bold', 'marginRight': '10px'}),
        dcc.Dropdown(
            id='table-selector',
            options=[{'label': table['title'], 'value': idx} for idx, table in enumerate(data)],
            value=0 if data else None,
            style={'width': '100%', 'maxWidth': '500px'}
        )
    ], style={'padding': '20px', 'backgroundColor': colors['card'], 'borderRadius': '10px',
              'marginBottom': '20px', 'boxShadow': '0 2px 4px rgba(0,0,0,0.1)'}),
    
    # Key Metrics Cards
    html.Div(id='metrics-cards', style={'marginBottom': '20px'}),
    
    # Charts Row 1
    html.Div([
        html.Div([
            dcc.Graph(id='traffic-trend-chart')
        ], style={'width': '100%', 'display': 'inline-block'}),
    ], style={'marginBottom': '20px'}),
    
    # Charts Row 2
    html.Div([
        html.Div([
            dcc.Graph(id='yoy-comparison-chart')
        ], style={'width': '48%', 'display': 'inline-block', 'marginRight': '2%'}),
        html.Div([
            dcc.Graph(id='mom-change-chart')
        ], style={'width': '48%', 'display': 'inline-block', 'marginLeft': '2%'}),
    ], style={'marginBottom': '20px'}),
    
    # Data Table and Summary
    html.Div([
        html.Div([
            html.H3("Monthly Data", style={'color': colors['text'], 'marginBottom': '15px'}),
            html.Div(id='data-table')
        ], style={'width': '48%', 'display': 'inline-block', 'marginRight': '2%',
                  'verticalAlign': 'top'}),
        html.Div([
            html.H3("AI Insights", style={'color': colors['text'], 'marginBottom': '15px'}),
            html.Div(id='summary-box', style={
                'backgroundColor': '#f8f9fa',
                'padding': '20px',
                'borderRadius': '8px',
                'border': '1px solid #dee2e6',
                'minHeight': '300px',
                'maxHeight': '500px',
                'overflowY': 'auto'
            })
        ], style={'width': '48%', 'display': 'inline-block', 'marginLeft': '2%',
                  'verticalAlign': 'top'}),
    ], style={'marginBottom': '20px'}),
    
], style={'backgroundColor': colors['background'], 'padding': '20px', 'fontFamily': 'Arial, sans-serif'})

@app.callback(
    [Output('metrics-cards', 'children'),
     Output('traffic-trend-chart', 'figure'),
     Output('yoy-comparison-chart', 'figure'),
     Output('mom-change-chart', 'figure'),
     Output('data-table', 'children'),
     Output('summary-box', 'children')],
    [Input('table-selector', 'value')]
)
def update_dashboard(selected_table_idx):
    if not data or selected_table_idx is None:
        empty_fig = go.Figure()
        empty_fig.add_annotation(text="No data available", showarrow=False)
        return [], empty_fig, empty_fig, empty_fig, html.Div("No data"), html.Div("No data")
    
    table_data = data[selected_table_idx]
    monthly_df = pd.DataFrame(table_data['monthly_data'])
    
    # Filter out rows with no data
    monthly_df = monthly_df[monthly_df['2024'].notna() | monthly_df['2025'].notna()]
    
    # Convert month names to month numbers for sorting
    month_order = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                   'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    monthly_df['month_num'] = monthly_df['month'].apply(
        lambda x: month_order.index(x[:3]) if x[:3] in month_order else 99
    )
    monthly_df = monthly_df.sort_values('month_num')
    
    # 1. Metrics Cards
    totals = table_data['totals']
    yoy_change = totals['yoy_total'] if totals['yoy_total'] is not None else 0
    
    metrics_cards = html.Div([
        html.Div([
            html.H4("Total 2024", style={'margin': '0', 'color': colors['text'], 'fontSize': '0.9rem'}),
            html.H2(f"{totals['2024']:,.0f}" if totals['2024'] else "N/A", 
                   style={'margin': '5px 0', 'color': colors['primary'], 'fontSize': '2rem'})
        ], style={'flex': '1', 'padding': '20px', 'backgroundColor': colors['card'],
                  'borderRadius': '8px', 'marginRight': '10px', 'boxShadow': '0 2px 4px rgba(0,0,0,0.1)'}),
        html.Div([
            html.H4("Total 2025", style={'margin': '0', 'color': colors['text'], 'fontSize': '0.9rem'}),
            html.H2(f"{totals['2025']:,.0f}" if totals['2025'] else "N/A",
                   style={'margin': '5px 0', 'color': colors['secondary'], 'fontSize': '2rem'})
        ], style={'flex': '1', 'padding': '20px', 'backgroundColor': colors['card'],
                  'borderRadius': '8px', 'marginRight': '10px', 'boxShadow': '0 2px 4px rgba(0,0,0,0.1)'}),
        html.Div([
            html.H4("YOY Change", style={'margin': '0', 'color': colors['text'], 'fontSize': '0.9rem'}),
            html.H2(f"{yoy_change:+.1f}%" if yoy_change is not None else "N/A",
                   style={'margin': '5px 0', 
                          'color': colors['accent'] if yoy_change < 0 else colors['secondary'],
                          'fontSize': '2rem'})
        ], style={'flex': '1', 'padding': '20px', 'backgroundColor': colors['card'],
                  'borderRadius': '8px', 'boxShadow': '0 2px 4px rgba(0,0,0,0.1)'}),
    ], style={'display': 'flex', 'gap': '10px'})
    
    # 2. Traffic Trend Chart
    trend_fig = go.Figure()
    
    if monthly_df['2023'].notna().any():
        trend_fig.add_trace(go.Scatter(
            x=monthly_df['month'],
            y=monthly_df['2023'],
            mode='lines+markers',
            name='2023',
            line=dict(color='#95a5a6', width=2),
            marker=dict(size=8)
        ))
    
    if monthly_df['2024'].notna().any():
        trend_fig.add_trace(go.Scatter(
            x=monthly_df['month'],
            y=monthly_df['2024'],
            mode='lines+markers',
            name='2024',
            line=dict(color='#3498db', width=3),
            marker=dict(size=10)
        ))
    
    if monthly_df['2025'].notna().any():
        trend_fig.add_trace(go.Scatter(
            x=monthly_df['month'],
            y=monthly_df['2025'],
            mode='lines+markers',
            name='2025',
            line=dict(color='#2ecc71', width=3),
            marker=dict(size=10)
        ))
    
    trend_fig.update_layout(
        title='Traffic Trends Over Time',
        xaxis_title='Month',
        yaxis_title='Sessions/Traffic',
        hovermode='x unified',
        plot_bgcolor='white',
        paper_bgcolor='white',
        font=dict(color=colors['text']),
        height=400
    )
    
    # 3. YOY Comparison Chart
    yoy_df = monthly_df[monthly_df['yoy'].notna()]
    yoy_fig = go.Figure()
    
    if not yoy_df.empty:
        colors_yoy = ['#e74c3c' if x < 0 else '#2ecc71' for x in yoy_df['yoy']]
        yoy_fig.add_trace(go.Bar(
            x=yoy_df['month'],
            y=yoy_df['yoy'],
            marker_color=colors_yoy,
            text=[f"{x:.1f}%" for x in yoy_df['yoy']],
            textposition='outside',
            name='YOY %'
        ))
    
    yoy_fig.update_layout(
        title='Year-over-Year Change by Month',
        xaxis_title='Month',
        yaxis_title='YOY Change (%)',
        plot_bgcolor='white',
        paper_bgcolor='white',
        font=dict(color=colors['text']),
        height=350
    )
    
    # 4. Month-over-Month Change Chart
    mom_df = monthly_df[monthly_df['lm'].notna()]
    mom_fig = go.Figure()
    
    if not mom_df.empty:
        colors_mom = ['#e74c3c' if x < 0 else '#2ecc71' for x in mom_df['lm']]
        mom_fig.add_trace(go.Bar(
            x=mom_df['month'],
            y=mom_df['lm'],
            marker_color=colors_mom,
            text=[f"{x:.1f}%" for x in mom_df['lm']],
            textposition='outside',
            name='MoM %'
        ))
    
    mom_fig.update_layout(
        title='Month-over-Month Change (2025)',
        xaxis_title='Month',
        yaxis_title='MoM Change (%)',
        plot_bgcolor='white',
        paper_bgcolor='white',
        font=dict(color=colors['text']),
        height=350
    )
    
    # 5. Data Table
    display_df = monthly_df[['month', '2023', '2024', '2025', 'yoy', 'lm']].copy()
    display_df.columns = ['Month', '2023', '2024', '2025', 'YOY %', 'MoM %']
    
    # Format numbers
    for col in ['2023', '2024', '2025']:
        display_df[col] = display_df[col].apply(lambda x: f"{x:,.0f}" if pd.notna(x) else "N/A")
    display_df['YOY %'] = display_df['YOY %'].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "N/A")
    display_df['MoM %'] = display_df['MoM %'].apply(lambda x: f"{x:.1f}%" if pd.notna(x) else "N/A")
    
    data_table = dash_table.DataTable(
        data=display_df.to_dict('records'),
        columns=[{"name": i, "id": i} for i in display_df.columns],
        style_cell={
            'textAlign': 'left',
            'padding': '10px',
            'fontFamily': 'Arial'
        },
        style_header={
            'backgroundColor': colors['primary'],
            'color': 'white',
            'fontWeight': 'bold'
        },
        style_data={
            'backgroundColor': 'white',
            'color': colors['text']
        },
        style_data_conditional=[
            {
                'if': {'row_index': 'odd'},
                'backgroundColor': '#f8f9fa'
            }
        ]
    )
    
    # 6. Summary Box
    summary_text = table_data.get('summary', 'No summary available')
    summary_content = html.Div([
        html.P(line, style={'margin': '10px 0', 'lineHeight': '1.6', 'color': colors['text']})
        for line in summary_text.split('\n') if line.strip()
    ])
    
    return metrics_cards, trend_fig, yoy_fig, mom_fig, data_table, summary_content

if __name__ == '__main__':
    print("Starting Traffic Insight Dashboard...")
    print(f"Loaded {len(data)} tables")
    print("Dashboard available at: http://127.0.0.1:8050")
    app.run(debug=True, host='127.0.0.1', port=8050)
