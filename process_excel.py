import openpyxl
from openpyxl import load_workbook
from openpyxl import styles
from openpyxl import utils
import requests
import json

# OpenRouter API configuration
API_KEY = "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"
API_URL = "https://openrouter.ai/api/v1/chat/completions"
MODEL = "tngtech/deepseek-r1t2-chimera:free"

def call_openrouter(prompt, is_summary=False):
    """Call OpenRouter API for calculations or summaries"""
    try:
        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json"
        }
        
        if is_summary:
            system_message = "You are an expert data analyst specializing in traffic and session data analysis. Provide insightful, well-structured summaries that highlight key trends, patterns, and performance metrics. Be thorough yet concise, covering growth trends, percentage changes, seasonal patterns, and notable insights."
            max_tokens = 700
            temperature = 0.3
        else:
            system_message = "You are a mathematical calculation assistant. Provide accurate numerical results only."
            max_tokens = 500
            temperature = 0.2
        
        payload = {
            "model": MODEL,
            "messages": [
                {
                    "role": "system",
                    "content": system_message
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            "temperature": temperature,
            "max_tokens": max_tokens
        }
        
        response = requests.post(API_URL, headers=headers, json=payload, timeout=60)
        response.raise_for_status()
        
        result = response.json()
        if 'choices' in result and len(result['choices']) > 0:
            content = result['choices'][0]['message']['content'].strip()
            
            if is_summary:
                return content
            else:
                # Try to extract numeric value from response
                try:
                    import re
                    # Look for number in the response (including decimals and negative)
                    numbers = re.findall(r'-?\d+\.?\d*', content)
                    if numbers:
                        return float(numbers[0])
                except:
                    pass
        return None
    except requests.exceptions.RequestException as e:
        print(f"  OpenRouter API call failed: {e}")
        return None
    except Exception as e:
        print(f"  Error processing OpenRouter response: {e}")
        return None

def safe_float(value):
    """Safely convert a value to float, handling formulas and strings"""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        # Skip if it's a header or formula string
        if value.startswith('=') or 'Year' in value or 'Sessions' in value:
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None
    return None

# Flag to control OpenRouter usage (set to False to skip API calls for faster processing)
USE_OPENROUTER = False  # Uses OpenRouter API for calculation validation

def calculate_with_openrouter(calculation_type, values):
    """Use OpenRouter to perform calculations with fallback to direct calculation"""
    # Perform direct calculation first (faster and more reliable)
    result = None
    
    if calculation_type == "total":
        result = sum([v for v in values if v is not None])
        # Use OpenRouter for validation on totals (important calculations)
        if USE_OPENROUTER and len(values) > 0 and len(values) <= 12:
            values_str = ", ".join([str(v) for v in values if v is not None])
            prompt = f"Calculate the sum of these numbers: {values_str}. Return only the numeric result."
            api_result = call_openrouter(prompt)
            if api_result is not None and abs(api_result - result) < 0.01:
                result = api_result
    
    elif calculation_type == "yoy_percent":
        """
        Calculate Year-Over-Year percentage change (2024 to 2025)
        
        Formula: YOY % = ((Sessions 2025 - Sessions 2024) / Sessions 2024) * 100
        
        Example: Jan 2024 = 5,288, Jan 2025 = 3,892
                 YOY % = ((3892 - 5288) / 5288) * 100 = -26.4%
        """
        val_2024, val_2025 = values
        val_2024 = safe_float(val_2024)
        val_2025 = safe_float(val_2025)
        if val_2024 is None or val_2025 is None or val_2024 == 0:
            return None
        
        # Calculate: ((2025 - 2024) / 2024) * 100
        result = ((val_2025 - val_2024) / val_2024) * 100
        
        # Use OpenRouter for validation
        if USE_OPENROUTER:
            prompt = f"Calculate year-over-year percentage change. Year 2024: {val_2024}, Year 2025: {val_2025}. Formula: ((2025 - 2024) / 2024) * 100. Return only the numeric result."
            api_result = call_openrouter(prompt)
            if api_result is not None and abs(api_result - result) < 1.0:  # Allow small difference
                result = api_result  # Use API result if close
    
    elif calculation_type == "lm_percent":
        """
        Calculate Last Month percentage change (month-over-month for 2025)
        
        Formula: LM % = ((Current Month - Previous Month) / Previous Month) * 100
        
        Example: Jan 2025 = 3,892, Feb 2025 = 3,977
                 LM % = ((3977 - 3892) / 3892) * 100 = 2.18%
        """
        current, previous = values
        current = safe_float(current)
        previous = safe_float(previous)
        if previous is None or current is None or previous == 0:
            return None
        
        # Calculate: ((current - previous) / previous) * 100
        result = ((current - previous) / previous) * 100
        
        # Use OpenRouter for validation (optional, can be disabled for speed)
        # if USE_OPENROUTER:
        #     prompt = f"Calculate month-over-month percentage change. Previous: {previous}, Current: {current}. Formula: ((current - previous) / previous) * 100. Return only the numeric result."
        #     api_result = call_openrouter(prompt)
        #     if api_result is not None and abs(api_result - result) < 1.0:
        #         result = api_result
    
    
    return result

def find_table_boundaries(ws, start_row):
    """Find the boundaries of a table starting at start_row"""
    # Header row is typically start_row or start_row + 1
    header_row = start_row
    if ws.cell(start_row, 2).value != 'Month':
        header_row = start_row + 1
    
    # Find data rows (Jan through Dec)
    data_start = header_row + 1
    data_end = data_start
    
    months = ['Jan', 'January', 'Feb', 'February', 'March', 'Mar', 'Apr', 'April', 
              'May', 'June', 'Jun', 'July', 'Jul', 'Aug', 'August', 
              'Sep', 'September', 'Oct', 'October', 'Nov', 'November', 'Dec', 'December']
    
    for i in range(data_start, min(data_start + 15, ws.max_row + 1)):
        month_val = ws.cell(i, 2).value
        if month_val:
            month_str = str(month_val).strip()
            # Check if it's a month (case-insensitive)
            if any(month.lower() in month_str.lower() or month_str.lower() in month.lower() for month in months):
                data_end = i
            elif month_str == 'Total':
                total_row = i
                break
    else:
        # Find Total row
        for i in range(data_end + 1, min(data_end + 5, ws.max_row + 1)):
            if ws.cell(i, 2).value == 'Total':
                total_row = i
                break
        else:
            total_row = data_end + 1
    
    # Find % Change row
    percent_change_row = total_row + 1
    if ws.cell(percent_change_row, 2).value != '% Change':
        percent_change_row = None
    
    return {
        'header_row': header_row,
        'data_start': data_start,
        'data_end': data_end,
        'total_row': total_row,
        'percent_change_row': percent_change_row
    }

def generate_table_summary(ws, title, boundaries, data_start, data_end, total_row):
    """Generate LLM summary for a table using OpenRouter"""
    try:
        # Collect table data
        table_data = []
        for row in range(data_start, data_end + 1):
            month = ws.cell(row, 2).value
            val_2023 = safe_float(ws.cell(row, 3).value)
            val_2024 = safe_float(ws.cell(row, 4).value)
            val_2025 = safe_float(ws.cell(row, 5).value)
            yoy = ws.cell(row, 6).value
            lm = ws.cell(row, 7).value
            
            table_data.append({
                'month': str(month) if month else '',
                '2023': val_2023,
                '2024': val_2024,
                '2025': val_2025,
                'yoy': yoy,
                'lm': lm
            })
        
        # Get totals
        total_2023 = safe_float(ws.cell(total_row, 3).value)
        total_2024 = safe_float(ws.cell(total_row, 4).value)
        total_2025 = safe_float(ws.cell(total_row, 5).value)
        yoy_total = ws.cell(total_row, 6).value
        
        # Build data summary for prompt
        data_summary = f"Table: {title}\n\n"
        data_summary += "Monthly Data:\n"
        for data in table_data:
            if data['2024'] is not None or data['2025'] is not None:
                line = f"{data['month']}: "
                if data['2024'] is not None:
                    line += f"2024={data['2024']:.0f}, "
                if data['2025'] is not None:
                    line += f"2025={data['2025']:.0f}"
                # yoy and lm are stored as decimals (0.264 = 26.4%), convert to percentage for display
                if data['yoy'] is not None:
                    yoy_pct = float(data['yoy']) * 100
                    line += f", YOY={yoy_pct:.2f}%"
                if data['lm'] is not None:
                    lm_pct = float(data['lm']) * 100
                    line += f", LM={lm_pct:.2f}%"
                data_summary += line + "\n"
        
        data_summary += f"\nTotals: "
        if total_2024 is not None:
            data_summary += f"2024={total_2024:.0f}, "
        if total_2025 is not None:
            data_summary += f"2025={total_2025:.0f}"
        if yoy_total is not None:
            yoy_total_pct = float(yoy_total) * 100
            data_summary += f", YOY Total={yoy_total_pct:.2f}%"
        
        # Create prompt
        prompt = f"""Analyze the following traffic/session data table and provide a comprehensive yet concise summary (8-10 sentences).

Provide detailed analysis covering:
1. Overall performance trends - summarize total volumes and year-over-year growth/decline patterns
2. Year-over-Year (YOY) analysis - highlight significant YOY percentage changes, identify best and worst performing months
3. Month-over-Month (LM) trends - analyze recent momentum, identify growth acceleration or deceleration patterns
4. Seasonal patterns - note any recurring monthly patterns or anomalies
5. Key insights - identify the most significant findings, explain potential drivers, and highlight actionable observations

{data_summary}

Structure your response with:
- Opening statement on overall performance
- Detailed breakdown of YOY trends with specific percentages
- Analysis of recent month-over-month momentum
- Identification of peak and low performance periods
- Closing insight on implications and trends
- Do not use Markdown formatting.
- Do not use **, *, _, `, or bullet symbols.
- Respond in plain text only.
Be thorough but concise, ensuring each sentence adds meaningful value to the analysis."""
        
        print(f"  → Generating LLM summary...")
        summary = call_openrouter(prompt, is_summary=True)
        
        if summary:
            formatted_summary = summary.replace("**", "")
            return formatted_summary
        else:
            return None
            
    except Exception as e:
        print(f"  ⚠ Error generating summary: {e}")
        return None

def calculate_table(ws, table_info):
    """Calculate and fill in all calculations for a table (totals, YOY%, LM%)"""
    title = table_info['title']
    start_row = table_info['row_idx']
    
    print(f"\nCalculating table: {title}")
    
    boundaries = find_table_boundaries(ws, start_row)
    header_row = boundaries['header_row']
    data_start = boundaries['data_start']
    data_end = boundaries['data_end']
    total_row = boundaries['total_row']
    percent_change_row = boundaries['percent_change_row']
    
    # Column indices (0-based, but openpyxl uses 1-based)
    # Column A=1, B=2 (Month), C=3 (2023), D=4 (2024), E=5 (2025), F=6 (YOY%), G=7 (LM%)
    
    # Get aligned 2025 and 2024 values for YOY total calculation
    values_2025_aligned = []
    values_2024_aligned = []
    for row in range(data_start, data_end + 1):
        val_2025 = safe_float(ws.cell(row, 5).value)  # Column E (2025)
        val_2024 = safe_float(ws.cell(row, 4).value)  # Column D (2024)
        if val_2025 is not None and val_2024 is not None:
            values_2025_aligned.append(val_2025)
            values_2024_aligned.append(val_2024)
    
    # Calculate and fill 2025 Total (aligned with 2024 where both years exist)
    if values_2025_aligned:
        total_2025 = calculate_with_openrouter("total", values_2025_aligned)
        ws.cell(total_row, 5).value = total_2025
    else:
        total_2025 = None
    
    # Calculate YOY % and LM% for each month
    prev_month_2025 = None
    for row in range(data_start, data_end + 1):
        month = ws.cell(row, 2).value
        val_2024 = safe_float(ws.cell(row, 4).value)  # Column D
        val_2025 = safe_float(ws.cell(row, 5).value)  # Column E
        
        # Calculate YOY % (2024-2025)
        if val_2024 is not None and val_2025 is not None:
            yoy_percent = calculate_with_openrouter("yoy_percent", [val_2024, val_2025])
            if yoy_percent is not None:
                cell = ws.cell(row, 6)
                cell.value = round(yoy_percent / 100, 4)
                cell.number_format = '0.00%'
        
        # Calculate LM% (2025) - month-over-month change
        if val_2025 is not None and prev_month_2025 is not None:
            lm_percent = calculate_with_openrouter("lm_percent", [val_2025, prev_month_2025])
            if lm_percent is not None:
                cell = ws.cell(row, 7)
                cell.value = round(lm_percent / 100, 4)
                cell.number_format = '0.00%'
        
        if val_2025 is not None:
            prev_month_2025 = val_2025
    
    # Calculate total_2024 from aligned values (for YOY total vs 2025)
    total_2024_aligned = calculate_with_openrouter("total", values_2024_aligned) if values_2024_aligned else None
    
    if total_2024_aligned is not None and total_2025 is not None:
        yoy_total = ((total_2025 - total_2024_aligned) / total_2024_aligned) * 100
        cell = ws.cell(total_row, 6)
        cell.value = round(yoy_total / 100, 4)
        cell.number_format = '0.00%'
    else:
        ws.cell(total_row, 6).value = None
    
    # Skip updating the '% Change' row in the Excel file if it exists.
    # This allows the workbook to keep its original formulas/values for that row.
    if percent_change_row:
        pass
    
    print(f"  ✓ Calculations completed for: {title}")

def add_table_summary(ws, table_info):
    """Generate and add LLM summary for a table using OpenRouter"""
    title = table_info['title']
    start_row = table_info['row_idx']
    
    print(f"\nGenerating summary for: {title}")
    
    boundaries = find_table_boundaries(ws, start_row)
    data_start = boundaries['data_start']
    data_end = boundaries['data_end']
    total_row = boundaries['total_row']
    
    # Generate LLM summary (starting at column H = 8, at data_start row)
    summary = generate_table_summary(ws, title, boundaries, data_start, data_end, total_row)
    if summary:
        # Write summary starting at column H (8), starting from data_start row
        summary_start_col = 8  # Column H
        summary_start_row = data_start
        
        # Split summary into lines and write to cells (max 10 rows to avoid overflow)
        summary_lines = summary.split('\n')[:10]
        for idx, line in enumerate(summary_lines):
            if line.strip():
                cell = ws.cell(summary_start_row + idx, summary_start_col)
                cell.value = line.strip()
                # Enable text wrapping
                cell.alignment = styles.Alignment(wrap_text=True, vertical='top')
        
        # Set column width for better readability
        ws.column_dimensions[utils.get_column_letter(summary_start_col)].width = 60
        
        print(f"  ✓ LLM summary added (column H, starting row {summary_start_row})")
    else:
        print(f"  ⚠ Failed to generate summary for: {title}")

def main():
    # Load workbook
    wb = load_workbook('keywords.xlsx')
    ws = wb['Traffic-Status']
    
    # Find all tables
    tables = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False), 1):
        cell_a = row[0]
        if cell_a.value and isinstance(cell_a.value, str):
            # Check if next row or current row has 'Month' in column B
            if row[1].value == 'Month' or (row_idx < ws.max_row and ws.cell(row_idx+1, 2).value == 'Month'):
                tables.append({
                    'title': cell_a.value,
                    'row_idx': row_idx
                })
    
    print(f"Found {len(tables)} tables to process")
    
    # ============================================
    # PASS 1: Calculate all tables (totals, YOY%, LM%)
    # ============================================
    print("\n" + "="*80)
    print("PASS 1: Filling all calculations (Totals, YOY%, LM%)")
    print("="*80)
    
    for table in tables:
        try:
            calculate_table(ws, table)
        except Exception as e:
            print(f"Error calculating table '{table['title']}': {e}")
            import traceback
            traceback.print_exc()
    
    # Save after calculations
    wb.save('keywords.xlsx')
    print("\n✓ All calculations completed and saved")
    
    # ============================================
    # PASS 2: Generate LLM summaries for all tables
    # ============================================
    print("\n" + "="*80)
    print("PASS 2: Generating LLM summaries for all tables")
    print("="*80)
    
    for table in tables:
        try:
            add_table_summary(ws, table)
        except Exception as e:
            print(f"Error generating summary for table '{table['title']}': {e}")
            import traceback
            traceback.print_exc()
    
    # Final save
    wb.save('keywords.xlsx')
    print("\n✓ Excel file updated successfully with all calculations and summaries!")

if __name__ == "__main__":
    main()
