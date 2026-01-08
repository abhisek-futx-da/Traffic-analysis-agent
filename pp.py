import openpyxl
from openpyxl import load_workbook
from openpyxl import styles
from openpyxl import utils
import requests
import json


# LangGraph additions ONLY

from langgraph.graph import StateGraph, END
from typing import TypedDict

# OpenRouter API configuration
API_KEY = "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"
API_URL = "https://openrouter.ai/api/v1/chat/completions"
MODEL = "tngtech/deepseek-r1t2-chimera:free"


def call_openrouter(prompt, is_summary=False):
    try:
        headers = {
            "Authorization": f"Bearer {API_KEY}",
            "Content-Type": "application/json"
        }

        if is_summary:
            system_message = (
                "You are an expert data analyst specializing in traffic and session data analysis. "
                "Provide insightful, well-structured summaries that highlight key trends, patterns, "
                "and performance metrics. Be thorough yet concise, covering growth trends, "
                "percentage changes, seasonal patterns, and notable insights."
            )
            max_tokens = 700
            temperature = 0.3
        else:
            system_message = "You are a mathematical calculation assistant. Provide accurate numerical results only."
            max_tokens = 500
            temperature = 0.2

        payload = {
            "model": MODEL,
            "messages": [
                {"role": "system", "content": system_message},
                {"role": "user", "content": prompt}
            ],
            "temperature": temperature,
            "max_tokens": max_tokens
        }

        response = requests.post(API_URL, headers=headers, json=payload, timeout=60)
        response.raise_for_status()

        result = response.json()
        if 'choices' in result and result['choices']:
            content = result['choices'][0]['message']['content'].strip()
            if is_summary:
                return content
            import re
            numbers = re.findall(r'-?\d+\.?\d*', content)
            if numbers:
                return float(numbers[0])
        return None

    except Exception as e:
        print(f"  OpenRouter error: {e}")
        return None


def safe_float(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        if value.startswith('=') or 'Year' in value or 'Sessions' in value:
            return None
        try:
            return float(value)
        except:
            return None
    return None


USE_OPENROUTER = False


def calculate_with_openrouter(calculation_type, values):
    if calculation_type == "total":
        return sum(v for v in values if v is not None)

    if calculation_type == "yoy_percent":
        v24, v25 = map(safe_float, values)
        if v24 in (None, 0) or v25 is None:
            return None
        return ((v25 - v24) / v24) * 100

    if calculation_type == "lm_percent":
        current, previous = map(safe_float, values)
        if previous in (None, 0) or current is None:
            return None
        return ((current - previous) / previous) * 100

    return None


def find_table_boundaries(ws, start_row):
    header_row = start_row if ws.cell(start_row, 2).value == 'Month' else start_row + 1
    data_start = header_row + 1
    data_end = data_start

    months = ['jan','feb','mar','apr','may','jun','jul','aug','sep','oct','nov','dec']

    for r in range(data_start, min(data_start + 15, ws.max_row + 1)):
        val = ws.cell(r, 2).value
        if val and any(m in str(val).lower() for m in months):
            data_end = r
        elif val == 'Total':
            break

    total_row = data_end + 1
    percent_change_row = total_row + 1 if ws.cell(total_row + 1, 2).value == '% Change' else None

    return {
        "header_row": header_row,
        "data_start": data_start,
        "data_end": data_end,
        "total_row": total_row,
        "percent_change_row": percent_change_row
    }


def generate_table_summary(ws, title, boundaries, data_start, data_end, total_row):
    data_lines = []

    for r in range(data_start, data_end + 1):
        month = ws.cell(r, 2).value
        v24 = safe_float(ws.cell(r, 4).value)
        v25 = safe_float(ws.cell(r, 5).value)
        if month and (v24 or v25):
            data_lines.append(f"{month}: 2024={v24}, 2025={v25}")

    summary_prompt = f"""
Analyze the following traffic/session data table and provide a concise summary.

Table: {title}
Data:
{chr(10).join(data_lines)}
"""

    return call_openrouter(summary_prompt, is_summary=True)


def calculate_table(ws, table_info):
    boundaries = find_table_boundaries(ws, table_info['row_idx'])
    prev_2025 = None

    for r in range(boundaries['data_start'], boundaries['data_end'] + 1):
        v24 = safe_float(ws.cell(r, 4).value)
        v25 = safe_float(ws.cell(r, 5).value)

        if v24 and v25:
            yoy = calculate_with_openrouter("yoy_percent", [v24, v25])
            ws.cell(r, 6).value = yoy / 100 if yoy is not None else None
            ws.cell(r, 6).number_format = '0.00%'

        if prev_2025 and v25:
            lm = calculate_with_openrouter("lm_percent", [v25, prev_2025])
            ws.cell(r, 7).value = lm / 100 if lm is not None else None
            ws.cell(r, 7).number_format = '0.00%'

        if v25:
            prev_2025 = v25


def add_table_summary(ws, table_info):
    boundaries = find_table_boundaries(ws, table_info['row_idx'])
    summary = generate_table_summary(
        ws,
        table_info['title'],
        boundaries,
        boundaries['data_start'],
        boundaries['data_end'],
        boundaries['total_row']
    )

    if summary:
        cell = ws.cell(boundaries['data_start'], 8)
        cell.value = summary
        cell.alignment = styles.Alignment(wrap_text=True)
        ws.column_dimensions[utils.get_column_letter(8)].width = 60


def main():
    wb = load_workbook('keywords.xlsx')
    ws = wb['Traffic-Status']

    tables = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=False), 1):
        if row[0].value and isinstance(row[0].value, str):
            if row[1].value == 'Month' or (row_idx < ws.max_row and ws.cell(row_idx+1, 2).value == 'Month'):
                tables.append({'title': row[0].value, 'row_idx': row_idx})

    print(f"Found {len(tables)} tables to process")

    print("\nPASS 1: Filling all calculations")
    for table in tables:
        calculate_table(ws, table)

    wb.save('keywords.xlsx')

    print("\nPASS 2: Generating summaries")
    for table in tables:
        add_table_summary(ws, table)

    wb.save('keywords.xlsx')
    print("\n✓ Done")



# LangGraph execution ONLY


class AppState(TypedDict):
    filename: str


def run_node(state: AppState) -> AppState:
    main()
    return state


def build_graph():
    graph = StateGraph(AppState)
    graph.add_node("run", run_node)
    graph.set_entry_point("run")
    graph.add_edge("run", END)
    return graph.compile()


if __name__ == "__main__":
    app = build_graph()
    app.invoke({"filename": "keywords.xlsx"})